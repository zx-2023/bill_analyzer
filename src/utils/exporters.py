"""Excel multi-sheet exporter for bill analysis data.

Generates an Excel workbook with formatted sheets for transactions,
category summaries, monthly trends, anomalies, and subscriptions.
"""
from io import BytesIO
from typing import Optional

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Column display names for each sheet
_TRANSACTION_COLUMNS = {
    "date": "交易时间",
    "platform": "平台",
    "counterparty": "交易对方",
    "description": "交易描述",
    "amount": "金额",
    "type": "收支类型",
    "category": "分类",
    "subcategory": "二级分类",
}

_CATEGORY_COLUMNS = {
    "category": "分类",
    "total_amount": "总金额",
    "count": "笔数",
    "percentage": "占比(%)",
}

_MONTHLY_COLUMNS = {
    "month": "月份",
    "total_expense": "总支出",
    "total_income": "总收入",
    "balance": "结余",
}

_ANOMALY_COLUMNS = {
    "date": "交易时间",
    "counterparty": "交易对方",
    "amount": "金额",
    "anomaly_type": "异常类型",
    "anomaly_score": "异常分数",
    "anomaly_reason": "异常原因",
}

_SUBSCRIPTION_COLUMNS = {
    "name": "订阅名称",
    "amount": "金额",
    "cycle": "周期",
    "count": "扣款次数",
}

# Header style constants
_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _select_and_rename(df: pd.DataFrame, column_map: dict) -> pd.DataFrame:
    """Select columns present in df and rename them to Chinese display names.

    Columns listed in column_map but missing from df are filled with empty strings
    so that every sheet always has the full set of headers.
    """
    result = pd.DataFrame()
    for src_col, display_name in column_map.items():
        if src_col in df.columns:
            result[display_name] = df[src_col]
        else:
            result[display_name] = ""
    return result


def _auto_fit_columns(ws, df: pd.DataFrame, min_width: int = 10, max_width: int = 40) -> None:
    """Set column widths based on header length and data content.

    Scans the header row and a sample of data rows to determine a reasonable
    width for each column.
    """
    for col_idx, col_name in enumerate(df.columns, start=1):
        # Start with the header length (CJK characters count as ~2 Latin chars)
        header_len = sum(2 if ord(c) > 127 else 1 for c in str(col_name))
        best_width = header_len + 4  # padding

        # Sample up to 100 data rows for max content width
        sample = df.iloc[:100, col_idx - 1]
        for val in sample:
            cell_len = sum(2 if ord(c) > 127 else 1 for c in str(val))
            if cell_len > best_width:
                best_width = cell_len

        clamped = max(min_width, min(best_width + 2, max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = clamped


def _style_header_row(ws) -> None:
    """Apply bold, colored styling to the first (header) row."""
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER


def _style_data_rows(ws, row_count: int) -> None:
    """Apply light borders and alignment to data rows."""
    data_alignment = Alignment(vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=row_count + 1):
        for cell in row:
            cell.border = _THIN_BORDER
            cell.alignment = data_alignment


def _write_sheet(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    column_map: dict,
    sheet_name: str,
) -> None:
    """Write a single sheet: select/rename columns, dump data, apply formatting."""
    prepared = _select_and_rename(df, column_map)

    prepared.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.sheets[sheet_name]
    _style_header_row(ws)
    _style_data_rows(ws, len(prepared))
    _auto_fit_columns(ws, prepared)

    # Freeze the header row so it stays visible while scrolling
    ws.freeze_panes = "A2"


def export_to_excel(
    transactions_df: pd.DataFrame,
    category_stats: pd.DataFrame,
    monthly_stats: pd.DataFrame,
    anomaly_df: Optional[pd.DataFrame] = None,
    subscription_df: Optional[pd.DataFrame] = None,
) -> bytes:
    """Export financial analysis data to a multi-sheet Excel workbook.

    Args:
        transactions_df: All transactions with columns matching Transaction.to_dict() keys.
        category_stats: Category summary with columns: category, total_amount, count, percentage.
        monthly_stats: Monthly trend with columns: month, total_expense, total_income, balance.
        anomaly_df: Anomaly transactions (optional, defaults to empty DataFrame).
        subscription_df: Subscription services (optional, defaults to empty DataFrame).

    Returns:
        bytes: Excel file content suitable for Streamlit st.download_button.
    """
    if anomaly_df is None:
        anomaly_df = pd.DataFrame()
    if subscription_df is None:
        subscription_df = pd.DataFrame()

    buf = BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        _write_sheet(writer, transactions_df, _TRANSACTION_COLUMNS, "交易明细")
        _write_sheet(writer, category_stats, _CATEGORY_COLUMNS, "分类汇总")
        _write_sheet(writer, monthly_stats, _MONTHLY_COLUMNS, "月度趋势")
        _write_sheet(writer, anomaly_df, _ANOMALY_COLUMNS, "异常交易")
        _write_sheet(writer, subscription_df, _SUBSCRIPTION_COLUMNS, "订阅服务")

    buf.seek(0)
    return buf.getvalue()
